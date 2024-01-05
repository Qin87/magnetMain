# external files
import os

import openpyxl
import pandas as pd
import pickle as pk
from datetime import datetime
import time

from openpyxl.reader.excel import load_workbook
from sklearn.metrics import balanced_accuracy_score, f1_score
import warnings

from layer.DGCN import SymModel
from layer.DiGCN import DiModel, DiGCN_IB
from nets_graphSHA.gat import create_gat
from nets_graphSHA.sage import create_sage
from utils.preprocess import F_in_out

warnings.filterwarnings("ignore")

# internal files
from gens_GraphSHA import sampling_idx_individual_dst, sampling_node_source, neighbor_sampling, \
    neighbor_sampling_BiEdge, neighbor_sampling_BiEdge_bidegree, neighbor_sampling_bidegreeOrigin
# from layer.DiGCN import *
from ArgsBen import parse_args
from utils.data_utils import make_longtailed_data_remove, get_idx_info, CrossEntropy, keep_all_data, \
    generate_masksRatio, get_step_split
from gens_GraphSHA import neighbor_sampling_bidegree, saliency_mixup, duplicate_neighbor, test_directed
from neighbor_dist import get_PPR_adj, get_heat_adj, get_ins_neighbor_dist
from nets_graphSHA.gcn import create_gcn
from utils.data_utils import get_dataset, load_directedData
from utils.Citation import *
from layer.geometric_baselines import *
from torch_geometric.utils import to_undirected
from utils.edge_data import get_appr_directed_adj, get_second_directed_adj


def main(args):
    # ********************* write to excel
    date_time = datetime.now().strftime('%m-%d-%H:%M')
    # date_time = datetime.now().strftime('%m-%d-%H:%M:%S')
    print(date_time)
    if args.IsDirectedData:
        excel_file_path = str(args.withAug) +str(args.AugDirect) +'Aug_' + date_time + '_'+ args.method_name + '_' + args.dataset.split('/')[
            0]+args.dataset.split('/')[
            1]  +'_dir.xlsx'
    else:
        excel_file_path = str(args.withAug) + str(args.AugDirect)+'Aug_' + args.method_name + '_' + args.undirect_dataset + date_time + '_undir.xlsx'
    print("excel_file_path is ", excel_file_path)

    writerBen = pd.ExcelWriter(excel_file_path, engine='openpyxl')    # a new excel file
    args_dict = vars(args)
    df = pd.DataFrame(args_dict.items(), columns=['Argument', 'Value'])
    combined_data = df
    if not combined_data.empty:
        combined_data.to_excel(writerBen, sheet_name="Args", index=False)
    else:
        print("DataFrame is empty. Not writing to the Excel file.")
    writerBen._save()
    # ********************

    if args.randomseed > 0:
        torch.manual_seed(args.randomseed)

    log_path = os.path.join(args.log_root, args.log_path, args.save_name, date_time)
    if os.path.isdir(log_path) is False:
        try:
            os.makedirs(log_path)
        except FileExistsError:
            print('Folder exists!')

    if args.IsDirectedData:
        dataset = load_directedData(args)
    else:
        path = args.data_path
        path = osp.join(path, args.undirect_dataset)
        dataset = get_dataset(args.undirect_dataset, path, split_type='full')
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    print("Dataset is ", dataset, "\nChosen from DirectedData: ", args.IsDirectedData)

    if os.path.isdir(log_path) is False:
        os.makedirs(log_path)

    data = dataset[0]
    data = data.to(device)

    global class_num_list, idx_info, prev_out, sample_times
    global data_train_maskOrigin, data_val_maskOrigin, data_test_maskOrigin  # data split: train, validation, test
    try:
        data.edge_weight = torch.FloatTensor(data.edge_weight)
    except:
        data.edge_weight = None

    if args.to_undirected:
        data.edge_index = to_undirected(data.edge_index)

    # copy GraphSHA
    if args.dataset.split('/')[0].startswith('dgl'):
        edges = torch.cat((data.edges()[0].unsqueeze(0), data.edges()[1].unsqueeze(0)), dim=0)
        data_y = data.ndata['label']
        data_train_maskOrigin, data_val_maskOrigin, data_test_maskOrigin = (
            data.ndata['train_mask'].clone(), data.ndata['val_mask'].clone(), data.ndata['test_mask'].clone())
        data_x = data.ndata['feat']
        dataset_num_features = data_x.shape[1]
    elif not args.IsDirectedData and args.undirect_dataset in ['Coauthor-CS', 'Amazon-Computers', 'Amazon-Photo']:
        edges = data.edge_index  # for torch_geometric librar
        data_y = data.y
        data_x = data.x
        dataset_num_features = dataset.num_features

        data_y = data_y.long()
        n_cls = (data_y.max() - data_y.min() + 1).cpu().numpy()
        n_cls = torch.tensor(n_cls).to(device)

        train_idx, valid_idx, test_idx, train_node = get_step_split(imb_ratio=args.imb_ratio,
                                                                    valid_each=int(data.x.shape[0] * 0.1 / n_cls),
                                                                    labeling_ratio=0.1,
                                                                    all_idx=[i for i in range(data.x.shape[0])],
                                                                    all_label=data.y.cpu().detach().numpy(),
                                                                    nclass=n_cls)

        data_train_maskOrigin = torch.zeros(data.x.shape[0]).bool().to(device)
        data_val_maskOrigin = torch.zeros(data.x.shape[0]).bool().to(device)
        data_test_maskOrigin = torch.zeros(data.x.shape[0]).bool().to(device)
        data_train_maskOrigin[train_idx] = True
        data_val_maskOrigin[valid_idx] = True
        data_test_maskOrigin[test_idx] = True
        train_idx = data_train_maskOrigin.nonzero().squeeze()
        train_edge_mask = torch.ones(data.edge_index.shape[1], dtype=torch.bool)

        class_num_list = [len(item) for item in train_node]
        idx_info = [torch.tensor(item) for item in train_node]
    elif dataset == 'Amazon-Photo':
        pass
    else:
        edges = data.edge_index  # for torch_geometric librar
        data_y = data.y
        data_train_maskOrigin, data_val_maskOrigin, data_test_maskOrigin = (data.train_mask.clone(), data.val_mask.clone(),data.test_mask.clone())
        data_x = data.x
        try:
            dataset_num_features = dataset.num_features
        except:
            dataset_num_features = data_x.shape[1]

    IsDirectedGraph = test_directed(edges)
    print("This is directed graph: ", IsDirectedGraph)
    print("data_x", data_x.shape)  # [11701, 300])

    data = data.to(device)

    data_y = data_y.long()
    n_cls = (data_y.max() - data_y.min() + 1).cpu().numpy()
    n_cls = torch.tensor(n_cls).to(device)

    criterion = CrossEntropy().to(device)

    try:
        splits = data_train_maskOrigin.shape[1]
        print("splits", splits)
        if len(data_test_maskOrigin.shape) == 1:
            data_test_maskOrigin = data_test_maskOrigin.unsqueeze(1).repeat(1, splits)
    except IndexError:
        splits = 1

    if args.method_name == 'APPNP' or 'DiG':
        edge_index1, edge_weights1 = get_appr_directed_adj(args.alpha, edges.long(), data_y.size(-1), data_x.dtype)
        edge_index1 = edge_index1.to(device)
        edge_weights1 = edge_weights1.to(device)
        if args.method_name[-2:] == 'ib':
            edge_index2, edge_weights2 = get_second_directed_adj(edges.long(), data_y.size(-1), data_x.dtype)
            edge_index2 = edge_index2.to(device)
            edge_weights2 = edge_weights2.to(device)
            SparseEdges = (edge_index1, edge_index2)
            edge_weight = (edge_weights1, edge_weights2)
            del edge_index2, edge_weights2
        else:
            SparseEdges = edge_index1
            edge_weight = edge_weights1
        del edge_index1, edge_weights1

    data = data.to(device)
    existing_data3 = pd.DataFrame()

    if args.method_name == 'GAT':
        model = create_gat(nfeat=dataset_num_features, nhid=args.feat_dim, nclass=n_cls, dropout=0.5,
                           nlayer=args.n_layer)  # SHA
    elif args.method_name == 'GCN':
        model = create_gcn(nfeat=dataset_num_features, nhid=args.feat_dim, nclass=n_cls, dropout=0.5,
                           nlayer=args.n_layer)  # SHA
    elif args.method_name == 'SAGE':
        model = create_sage(nfeat=dataset_num_features, nhid=args.feat_dim, nclass=n_cls, dropout=0.5,
                            nlayer=args.n_layer)
    elif args.method_name == 'GIN':
        model = GIN_ModelBen(data_x.size(-1), n_cls, filter_num=args.num_filter,
                             dropout=args.dropout, layer=args.layer).to(device)
    elif args.method_name == 'Cheb':
        model = ChebModelBen(data_x.size(-1), n_cls, K=args.K,
                             filter_num=args.num_filter, dropout=args.dropout,
                             layer=args.layer).to(device)
    elif args.method_name == 'APPNP':
        model = APPNP_ModelBen(data_x.size(-1), n_cls,
                               filter_num=args.num_filter, alpha=args.alpha,
                               dropout=args.dropout, layer=args.layer).to(device)
    elif args.method_name == 'DiG':
        if not args.method_name[-2:] == 'ib':
            model = DiModel(data_x.size(-1), n_cls, filter_num=args.num_filter,
                            dropout=args.dropout, layer=args.layer).to(device)
        else:
            model = DiGCN_IB(data_x.size(-1), hidden=args.num_filter,
                             n_cls=n_cls, dropout=args.dropout,
                             layer=args.layer).to(device)

    elif args.method_name == 'SymDiGCN':
        model = SymModel(data_x.size(-1), n_cls, filter_num=args.num_filter,
                         dropout=args.dropout, layer=args.layer).to(device)
    else:
        raise NotImplementedError
    try:
        print(model)  # # StandGCN2((conv1): GCNConv(3703, 64)  (conv2): GCNConv(64, 6))
    except:
        pass
    model.to(device)

    for split in range(splits):
        # if split >4:
        #     continue
        print("Beginning for split: ", split, datetime.now().strftime('%d-%H:%M:%S'))
        if splits == 1:
            data_train_mask, data_val_mask, data_test_mask = (data_train_maskOrigin.clone(),
                                                              data_val_maskOrigin.clone(),
                                                              data_test_maskOrigin.clone())
        else:
            try:
                data_train_mask, data_val_mask, data_test_mask = (data_train_maskOrigin[:, split].clone(),
                                                              data_val_maskOrigin[:, split].clone(),
                                                              data_test_maskOrigin[:, split].clone())
            except IndexError:
                print("testIndex ,", data_test_mask.shape, data_train_mask.shape, data_val_mask.shape)
                data_train_mask, data_val_mask = (data_train_maskOrigin[:, split].clone(),data_val_maskOrigin[:, split].clone())
                data_test_mask = data_test_maskOrigin[:, 1].clone()

        if args.CustomizeMask:
            data_train_mask, data_val_mask, data_test_mask = generate_masksRatio(data_y, TrainRatio=0.3, ValRatio=0.3)
        stats = data_y[data_train_mask]  # this is selected y. only train nodes of y
        n_data = []  # num of train in each class
        for i in range(n_cls):
            data_num = (stats == i).sum()
            n_data.append(int(data_num.item()))
        idx_info = get_idx_info(data_y, n_cls, data_train_mask)  # torch: all train nodes for each class
        if args.MakeImbalance:
            class_num_list, data_train_mask, idx_info, train_node_mask, train_edge_mask = \
                make_longtailed_data_remove(edges, data_y, n_data, n_cls, args.imb_ratio, data_train_mask.clone())
        else:
            class_num_list, data_train_mask, idx_info, train_node_mask, train_edge_mask = \
                keep_all_data(edges, data_y, n_data, n_cls, args.imb_ratio, data_train_mask)

        train_idx = data_train_mask.nonzero().squeeze()  # get the index of training data
        labels_local = data_y.view([-1])[train_idx]  # view([-1]) is "flattening" the tensor.
        train_idx_list = train_idx.cpu().tolist()
        local2global = {i: train_idx_list[i] for i in range(len(train_idx_list))}
        global2local = dict([val, key] for key, val in local2global.items())
        idx_info_list = [item.cpu().tolist() for item in idx_info]  # list of all train nodes for each class
        idx_info_local = [torch.tensor(list(map(global2local.get, cls_idx))) for cls_idx in
                          idx_info_list]  # train nodes position inside train

        if args.gdc == 'ppr':
            neighbor_dist_list = get_PPR_adj(data_x, edges[:, train_edge_mask], alpha=0.05, k=128, eps=None)
        elif args.gdc == 'hk':
            neighbor_dist_list = get_heat_adj(data_x, edges[:, train_edge_mask], t=5.0, k=None, eps=0.0001)
        elif args.gdc == 'none':
            neighbor_dist_list = get_ins_neighbor_dist(data_y.size(0), edges[:, train_edge_mask], data_train_mask,
                                                       device)

        #     #################################
        #     # Train/Validation/Test
        #     #################################
        test_accSHA = test_bacc = test_f1 = 0.0

        # from GraphSHA
        best_val_acc_f1 = 0
        saliency, prev_out = None, None

        existing_data2 = pd.DataFrame()
        CountNotImproved = 0
        for epoch in range(args.epochs):
            if CountNotImproved > 50:
                opt = torch.optim.Adam(model.parameters(), lr=10 * args.lr, weight_decay=args.l2)  # less accuracy*
            else:
                opt = torch.optim.Adam(model.parameters(), lr=args.lr, weight_decay=args.l2)  # less accuracy
            # opt = torch.optim.Adam(
            #     [dict(params=model.reg_params, weight_decay=5e-4), dict(params=model.non_reg_params, weight_decay=0), ],
            #     lr=args.lr)  # from SHA
            scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(opt, mode='min', factor=0.5, patience=100,
                                                                   verbose=False)
            start_time = time.time()
            model.train()
            model.to(device)
            opt.zero_grad()  # clear the gradients of the model's parameters.

            if args.withAug:
                if epoch > args.warmup:
                    prev_out_local = prev_out[train_idx]
                    sampling_src_idx, sampling_dst_idx = sampling_node_source(class_num_list, prev_out_local,
                                                                              idx_info_local, train_idx, args.tau,
                                                                              args.max, args.no_mask)
                    if args.AugDirect == 1:
                        new_edge_index = neighbor_sampling(data_x.size(0), edges[:, train_edge_mask], sampling_src_idx,
                                                           neighbor_dist_list)
                    elif args.AugDirect == 2:
                        new_edge_index = neighbor_sampling_BiEdge(data_x.size(0), edges[:, train_edge_mask],
                                                                  sampling_src_idx, neighbor_dist_list)
                    elif args.AugDirect == 4:
                        new_edge_index = neighbor_sampling_BiEdge_bidegree(data_x.size(0), edges[:, train_edge_mask],
                                                                           sampling_src_idx, neighbor_dist_list)
                    elif args.AugDirect == 20:
                        new_edge_index = neighbor_sampling_bidegree(data_x.size(0), edges[:, train_edge_mask],
                                                                    sampling_src_idx,
                                                                    neighbor_dist_list)  # has two types
                    elif args.AugDirect == 21:
                        new_edge_index = neighbor_sampling_bidegreeOrigin(data_x.size(0), edges[:, train_edge_mask],
                                                                          sampling_src_idx, neighbor_dist_list)

                    else:
                        pass
                    beta = torch.distributions.beta.Beta(1, 100)
                    lam = beta.sample((len(sampling_src_idx),)).unsqueeze(1)
                    new_x = saliency_mixup(data_x, sampling_src_idx, sampling_dst_idx, lam)

                    add_num = new_x.shape[0] - data_x.shape[0]  # Ben
                    new_train_mask = torch.ones(add_num, dtype=torch.bool, device=data_x.device)
                    new_train_mask = torch.cat((data_train_mask, new_train_mask), dim=0)  # add some train nodes
                    _new_y = data_y[sampling_src_idx.long()].clone()
                    new_y = torch.cat((data_y, _new_y), dim=0)  #

                    # get edge_weight
                    edge_index1, edge_weights1 = get_appr_directed_adj(args.alpha, new_edge_index.long(),
                                                                       new_y.size(-1), new_x.dtype)
                    edge_index1 = edge_index1.to(device)
                    edge_weights1 = edge_weights1.to(device)
                    if args.method_name[-2:] == 'ib':
                        edge_index2, edge_weights2 = get_second_directed_adj(new_edge_index.long(), new_y.size(-1),
                                                                             new_x.dtype)
                        edge_index2 = edge_index2.to(device)
                        edge_weights2 = edge_weights2.to(device)
                        new_SparseEdges = (edge_index1, edge_index2)
                        new_edge_weight = (edge_weights1, edge_weights2)
                        del edge_index2, edge_weights2
                    else:
                        new_SparseEdges = edge_index1
                        new_edge_weight = edge_weights1
                    # print("edge_weight", new_edge_weight.shape, data.y.shape)
                    del edge_index1, edge_weights1

                else:
                    sampling_src_idx, sampling_dst_idx = sampling_idx_individual_dst(class_num_list, idx_info, device)
                    beta = torch.distributions.beta.Beta(2, 2)
                    lam = beta.sample((len(sampling_src_idx),)).unsqueeze(1)
                    new_edge_index = duplicate_neighbor(data_x.size(0), edges[:, train_edge_mask], sampling_src_idx)
                    new_x = saliency_mixup(data_x, sampling_src_idx, sampling_dst_idx, lam)

                    add_num = len(sampling_src_idx)  # Ben
                    new_train_mask = torch.ones(add_num, dtype=torch.bool, device=data_x.device)
                    new_train_mask = torch.cat((data_train_mask, new_train_mask), dim=0)  # add some train nodes
                    _new_y = data_y[sampling_src_idx].clone()
                    new_y = torch.cat((data_y, _new_y), dim=0)  #
                    edge_index1, edge_weights1 = get_appr_directed_adj(args.alpha, new_edge_index.long(),
                                                                       new_y.size(-1), new_x.dtype)
                    edge_index1 = edge_index1.to(device)
                    edge_weights1 = edge_weights1.to(device)
                    if args.method_name[-2:] == 'ib':
                        edge_index2, edge_weights2 = get_second_directed_adj(new_edge_index.long(), new_y.size(-1),new_x.dtype)
                        edge_index2 = edge_index2.to(device)
                        edge_weights2 = edge_weights2.to(device)
                        new_SparseEdges = (edge_index1, edge_index2)
                        new_edge_weight = (edge_weights1, edge_weights2)
                        del edge_index2, edge_weights2
                    else:
                        new_SparseEdges = edge_index1
                        new_edge_weight = edge_weights1
                    del edge_index1, edge_weights1
                if args.method_name == 'SymDiGCN':
                    data.edge_index, edge_in, in_weight, edge_out, out_weight = F_in_out(edges,
                                                                                         data_y.size(-1),
                                                                                         data.edge_weight)
                    out = model(new_x, data.edge_index, edge_in, in_weight, edge_out, out_weight)
                else:
                    try:
                        out = model(new_x, new_SparseEdges, new_edge_weight)  #
                    except RuntimeError:
                        model.to('cpu')
                        new_x = new_x.to('cpu')
                        new_SparseEdges = new_SparseEdges.to('cpu')
                        new_edge_weight = new_edge_weight.to('cpu')
                        out = model(new_x, new_SparseEdges, new_edge_weight)  #
                        model.to(device)
                        new_x = new_x.to(device)
                        new_SparseEdges = new_SparseEdges.to(device)
                        new_edge_weight = new_edge_weight.to(device)
                    except TypeError:
                        try:
                            out = model(new_x, new_edge_index)
                        except TypeError:
                            out= model(new_x)

                prev_out = (out[:data_x.size(0)]).clone().to(device)
                _new_y = data_y[sampling_src_idx.long()].clone()    # AttributeError: 'tuple' object has no attribute 'detach'
                new_y = torch.cat((data_y[data_train_mask], _new_y), dim=0)
                new_y = new_y.to(out.device)
                criterion(out[new_train_mask], new_y).backward()

            else:  # # without aug
                if args.method_name == 'SymDiGCN':
                    data.edge_index, edge_in, in_weight, edge_out, out_weight = F_in_out(edges,
                                                                                         data_y.size(-1),
                                                                                         data.edge_weight)
                    try:
                        out = model(data_x, edges, edge_in, in_weight, edge_out, out_weight)
                    except:
                        model.to('cpu')
                        data_x = data_x.to('cpu')
                        edges = edges.to('cpu')
                        edge_in = edge_in.to('cpu')
                        edge_out = edge_out.to('cpu')
                        in_weight = in_weight.to('cpu')
                        out_weight = out_weight.to('cpu')
                        out = model(data_x, edges, edge_in, in_weight, edge_out, out_weight)
                        model.to(device)
                        data_x = data_x.to(device)
                        edges = edges.to(device)
                        edge_in = edge_in.to(device)
                        edge_out = edge_out.to(device)
                        in_weight = in_weight.to(device)
                        out_weight = out_weight.to(device)

                elif args.method_name == 'DiG':
                    out = model(data_x, SparseEdges, edge_weight)
                else:
                    out = model(data_x, edges)
                criterion(out[data_train_mask], data_y[data_train_mask]).backward()
            torch.cuda.empty_cache()

            with torch.no_grad():
                model.eval()
                if args.method_name == 'SymDiGCN':
                    out = model(data_x, edges[:, train_edge_mask], edge_in, in_weight, edge_out, out_weight)
                    # print(model)
                elif args.method_name == 'DiG':
                    out = model(data_x, SparseEdges, edge_weight)
                else:
                    out = model(data_x, edges[:, train_edge_mask])

                val_loss = F.cross_entropy(out[data_val_mask], data_y[data_val_mask])
            opt.step()
            scheduler.step(val_loss)
            # from graphSHA
            model.eval()
            if args.method_name == 'SymDiGCN':
                logits = model(data_x, edges[:, train_edge_mask], edge_in, in_weight, edge_out, out_weight)
            elif args.method_name == 'DiG':
                logits = model(data_x, SparseEdges, edge_weight)
            else:
                logits = model(data_x, edges[:, train_edge_mask])

            accs, baccs, f1s = [], [], []
            for mask in [data_train_mask, data_val_mask, data_test_mask]:
                # print(mask.shape, logits.shape)  # torch.Size([3327, 1]) torch.Size([3327, 6])
                pred = logits[mask].max(1)[1]
                y_pred = pred.cpu().numpy()
                y_true = data_y[mask].cpu().numpy()
                acc_epoch = pred.eq(data_y[mask]).sum().item() / mask.sum().item()
                bacc = balanced_accuracy_score(y_true, y_pred)
                f1 = f1_score(y_true, y_pred, average='macro')
                accs.append(acc_epoch)
                baccs.append(bacc)
                f1s.append(f1)
            train_accSHA, val_accSHA, tmp_test_acc = accs
            train_f1, val_f1, tmp_test_f1 = f1s
            val_acc_f1 = (val_accSHA + val_f1) / 2.
            if val_acc_f1 > best_val_acc_f1:
                best_val_acc_f1 = val_acc_f1
                test_accSHA = accs[2]
                test_bacc = baccs[2]
                test_f1 = f1s[2]
            else:
                CountNotImproved += 1
            if CountNotImproved > 500:
                print("Early stop at epoch: ", epoch)
                break

            end_time = time.time()
            epoch_time = end_time - start_time
            if epoch%2:
                Epoch_output_str = 'Epoch:{:3d}, Val_loss:{:6.2f}, time:{:6.2f}, test_Acc: {:6.2f}, test_bacc: {:6.2f}, test_f1: {:6.2f}'.format(epoch,val_loss,epoch_time, test_accSHA * 100, test_bacc * 100,test_f1 * 100)
            else:
                end_time_datetime = datetime.fromtimestamp(end_time)
                Epoch_output_str = 'Epoch:{:3d}, Val_loss:{:6.2f}, present time:{:s}, test_Acc: {:6.2f}, test_bacc: {:6.2f}, test_f1: {:6.2f}'.format(
                    epoch, val_loss, end_time_datetime.strftime("%H:%M:%S"), test_accSHA * 100, test_bacc * 100,
                    test_f1 * 100)
            df2 = pd.DataFrame({'Epoch_Output': [Epoch_output_str]})
            df2 = pd.concat([df2, existing_data2])
            existing_data2 = df2
            os.chdir(os.path.dirname(os.path.abspath(__file__)))
            try:
                workbook = openpyxl.load_workbook(excel_file_path)
            except:
                current_directory = os.getcwd()
                parent_directory = os.path.dirname(current_directory)
                os.chdir(parent_directory)
                workbook = openpyxl.load_workbook(excel_file_path)

            if 'Epoch'+str(split) in workbook.sheetnames:
                workbook.remove(workbook['Epoch'+str(split)])
                workbook.save(excel_file_path)
            workbook.close()
            writerBen = pd.ExcelWriter(excel_file_path, mode="a", engine="openpyxl")
            df2.to_excel(writerBen, sheet_name="Epoch"+str(split), index=False)
            writerBen._save()
        print('split: {:3d}, val_loss: {:6.2f}, test_Acc: {:6.2f}, test_bacc: {:6.2f}, test_f1: {:6.2f}'.format(split,val_loss, test_accSHA * 100, test_bacc * 100,test_f1 * 100))
        Split_output_str = 'split: {:3d}, val_loss: {:6.2f}, test_Acc: {:6.2f}, test_bacc: {:6.2f}, test_f1: {:6.2f}'.format(split,val_loss, test_accSHA * 100, test_bacc * 100,test_f1 * 100)
        df3 = pd.DataFrame({'Split_Output': [Split_output_str]})
        df3 = pd.concat([df3, existing_data3])
        existing_data3 = df3
        os.chdir(os.path.dirname(os.path.abspath(__file__)))
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
        except:
            current_directory = os.getcwd()
            parent_directory = os.path.dirname(current_directory)
            os.chdir(parent_directory)
            workbook = openpyxl.load_workbook(excel_file_path)
        if 'Split' in workbook.sheetnames:
            workbook.remove(workbook['Split'])
            workbook.save(excel_file_path)
        workbook.close()
        writerBen = pd.ExcelWriter(excel_file_path, mode="a", engine="openpyxl")
        df3.to_excel(writerBen, sheet_name="Split", index=False)
        writerBen._save()

    end_time = datetime.now().strftime('%m-%d-%H:%M:%S')
    time_str = 'startTime:'+ date_time + '\tend_time: ' + end_time
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    df4= pd.DataFrame({'Usedtime': [time_str]})
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except:
        current_directory = os.getcwd()
        parent_directory = os.path.dirname(current_directory)
        os.chdir(parent_directory)
        workbook = openpyxl.load_workbook(excel_file_path)
    workbook.close()
    writerBen = pd.ExcelWriter(excel_file_path, mode="a", engine="openpyxl")
    df4.to_excel(writerBen, sheet_name="Usedtime", index=False)
    writerBen._save()


if __name__ == "__main__":
    start_sum_time = time.time()
    args = parse_args()
    print(args)

    cuda_device = args.GPUdevice
    if torch.cuda.is_available():
        print("CUDA Device Index:", cuda_device)
        device = torch.device("cuda:%d" % cuda_device)
    else:
        print("CUDA is not available, using CPU.")
        device = torch.device("cpu")

    if args.debug:
        args.epochs = 1
    if args.dataset[:3] == 'syn':
        if args.dataset[4:7] == 'syn':
            if args.p_q not in [-0.08, -0.05]:
                args.dataset = 'syn/syn' + str(int(100 * args.p_q)) + 'Seed' + str(args.seed)
            elif args.p_q == -0.08:
                args.p_inter = -args.p_q
                args.dataset = 'syn/syn2Seed' + str(args.seed)
            elif args.p_q == -0.05:
                args.p_inter = -args.p_q
                args.dataset = 'syn/syn3Seed' + str(args.seed)
        elif args.dataset[4:10] == 'cyclic':
            args.dataset = 'syn/cyclic' + str(int(100 * args.p_q)) + 'Seed' + str(args.seed)
        else:
            args.dataset = 'syn/fill' + str(int(100 * args.p_q)) + 'Seed' + str(args.seed)
    dir_name = os.path.join(os.path.dirname(os.path.realpath(
        __file__)), '../result_arrays', args.log_path, args.dataset + '/')
    args.log_path = os.path.join(args.log_path, args.method_name, args.dataset)
    if not args.new_setting:
        if args.dataset[:3] == 'syn':
            if args.dataset[4:7] == 'syn':
                setting_dict = pk.load(open('syn_settings.pk', 'rb'))
                dataset_name_dict = {
                    0.95: 1, 0.9: 4, 0.85: 5, 0.8: 6, 0.75: 7, 0.7: 8, 0.65: 9, 0.6: 10
                }
                if args.p_inter == 0.1:
                    dataset = 'syn/syn' + str(dataset_name_dict[args.p_q])
                elif args.p_inter == 0.08:
                    dataset = 'syn/syn2'
                elif args.p_inter == 0.05:
                    dataset = 'syn/syn3'
                else:
                    raise ValueError('Please input the correct p_q and p_inter values!')
            elif args.dataset[4:10] == 'cyclic':
                setting_dict = pk.load(open('./Cyclic_setting_dict.pk', 'rb'))
                dataset_name_dict = {
                    0.95: 0, 0.9: 1, 0.85: 2, 0.8: 3, 0.75: 4, 0.7: 5, 0.65: 6
                }
                dataset = 'syn/syn_tri_' + str(dataset_name_dict[args.p_q])
            else:
                setting_dict = pk.load(open('./Cyclic_fill_setting_dict.pk', 'rb'))
                dataset_name_dict = {
                    0.95: 0, 0.9: 1, 0.85: 2, 0.8: 3
                }
                dataset = 'syn/syn_tri_' + str(dataset_name_dict[args.p_q]) + '_fill'
            setting_dict_curr = setting_dict[dataset][args.method_name].split(',')
            args.alpha = float(setting_dict_curr[setting_dict_curr.index('alpha') + 1])
            args.heads = int(setting_dict_curr[setting_dict_curr.index('heads') + 1])
            args.to_undirected = (setting_dict_curr[setting_dict_curr.index('to_undirected') + 1] == 'True')
            try:
                args.num_filter = int(setting_dict_curr[setting_dict_curr.index('num_filter') + 1])
            except ValueError:
                try:
                    args.num_filter = int(setting_dict_curr[setting_dict_curr.index('num_filters') + 1])
                except ValueError:
                    pass
            args.lr = float(setting_dict_curr[setting_dict_curr.index('lr') + 1])
            try:
                args.layer = int(setting_dict_curr[setting_dict_curr.index('layer') + 1])
            except ValueError:
                pass

    if os.path.isdir(dir_name) is False:
        try:
            os.makedirs(dir_name)
        except FileExistsError:
            print('Folder exists!')

    if args.method_name == 'GAT':
        save_name = args.method_name + 'lr' + str(int(args.lr * 1000)) + 'num_filters' + str(
            int(args.num_filter)) + 'tud' + str(args.to_undirected) + 'heads' + str(int(args.heads)) + 'layer' + str(
            int(args.layer))
    else:
        save_name = args.method_name + 'lr' + str(int(args.lr * 1000)) + 'num_filters' + str(
            int(args.num_filter)) + 'alpha' + str(int(100 * args.alpha)) + 'layer' + str(
            int(args.layer))  # Digraph and GCN
    args.save_name = save_name
    main(args)
    end_sum_time = time.time()
    total_time = end_sum_time- start_sum_time
    print("Total time(all splits): ", total_time)

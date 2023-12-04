# import pandas as pd
#
# def print_and_save_output(split, test_accSHA, test_bacc, test_f1, device):
#     out_str = 'split: {}, test_Acc: {:.2f}, test_bacc: {:.2f}, test_f1: {:.2f}'.format(split, test_accSHA * 100, test_bacc * 100, test_f1 * 100)
#     Eout_str = 'split: {}, test_Acc: {:.2f}, test_bacc: {:.2f}, test_f1: {:.2f}'.format(split, test_accSHA * 666666100, test_bacc * 100, test_f1 * 100)
#
#     # Print to console
#     print(out_str)
#     print('Device:', device)
#
#     # Create a DataFrame
#     df = pd.DataFrame({'Epoch_Output': [Eout_str],'Output': [out_str], 'Device': [device]})
#
#     # Write the DataFrame to an Excel file
#     df.to_excel('output.xlsx', index=False)
#
# # Example usage
# split = 1
# test_accSHA = 0.85
# test_bacc = 0.90
# test_f1 = 0.87
# device = 'cuda:0'
#
# print_and_save_output(split, test_accSHA, test_bacc, test_f1, device)
import os

import openpyxl
import pandas as pd

# # Assuming excel_file_path is defined elsewhere
# excel_file_path = 'your_excel_file.xlsx'
#
# # Dummy values for demonstration
# epoch = 1
# test_accSHA = 0.85
# test_bacc = 0.75
# test_f1 = 0.80
# epoch_time = 10.5
# split = 1
#
# # Construct the first DataFrame for the Epoch output
# epoch_output_str = 'Epoch:{}, time:{:2f}, test_Acc: {:.2f}, test_bacc: {:.2f}, test_f1: {:.2f}'.format(epoch, epoch_time, test_accSHA * 100, test_bacc * 100, test_f1 * 100)
# epoch_df = pd.DataFrame({'Epoch_Output': [epoch_output_str]})
#
# # Try to read the existing Excel file
# try:
#     existing_data = pd.read_excel(excel_file_path, sheet_name='Sheet1', engine='openpyxl')
#     combined_data = pd.concat([epoch_df, existing_data], ignore_index=True)
# except FileNotFoundError:
#     combined_data = epoch_df
#
# # Write the combined data to the Excel file with the first sheet
# combined_data.to_excel(excel_file_path, sheet_name='Sheet1', index=False, engine='openpyxl')
#
# # Construct the second DataFrame for the Split output
# split_output_str = 'split: {}, test_Acc: {:.2f}, test_bacc: {:.2f}, test_f1: {:.2f}'.format(split, test_accSHA * 100, test_bacc * 100, test_f1 * 100)
# split_df = pd.DataFrame({'Split_Output': [split_output_str]})
#
# # Try to read the existing Excel file again
# try:
#     existing_data = pd.read_excel(excel_file_path, sheet_name='Sheet2', engine='openpyxl')
#     combined_data = pd.concat([split_df, existing_data], ignore_index=True)
# except FileNotFoundError:
#     combined_data = split_df
#
# # Write the combined data to the Excel file with the second sheet
# combined_data.to_excel(excel_file_path, sheet_name='Sheet2', index=False, engine='openpyxl')
# import pandas as pd
#
# # Create a Pandas DataFrame for your data
# data1 = {'Column1': [1, 2, 3], 'Column2': ['A', 'B', 'C']}
# df1 = pd.DataFrame(data1)
#
# data2 = {'Column3': ['X', 'Y', 'Z'], 'Column4': [4, 5, 6]}
# df2 = pd.DataFrame(data2)
#
# # Specify the Excel file name
# excel_file = 'example.xlsx'
#
# # Write DataFrame df1 to Sheet1
# with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
#     df1.to_excel(writer, sheet_name='Sheet1', index=False)
#
# # Write DataFrame df2 to Sheet2
# with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
#     df2.to_excel(writer, sheet_name='Sheet2', index=False)
#
# # To write a new set of contents to the sheets again, you can repeat the process
# # with new DataFrames and the same ExcelWriter object.
#
# # For example:
# data3 = {'Column5': [7, 8, 9], 'Column6': ['P', 'Q', 'R']}
# df3 = pd.DataFrame(data3)
#
# data4 = {'Column7': ['M', 'N', 'O'], 'Column8': [10, 11, 12]}
# df4 = pd.DataFrame(data4)
#
# # Write new DataFrame df3 to Sheet1
# with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
#     df3.to_excel(writer, sheet_name='Sheet1', index=False)
#
# # Write new DataFrame df4 to Sheet2
# with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
#     df4.to_excel(writer, sheet_name='Sheet2', index=False)


print(os.getcwd())
filename = 'myfile.xlsx'
writerBen = pd.ExcelWriter(filename, engine="openpyxl")
df = pd.DataFrame([[1, 2, 3], [4, 5, 6]], columns=['A', 'B', 'C'])
df2 = pd.DataFrame([[11, 23, 33], [433,333, 622]], columns=['A', 'B', 'C'])
df.to_excel(writerBen,sheet_name='Sheet1')
# file = pd.ExcelFile(filename)
writerBen._save()   # close and save

writerBen = pd.ExcelWriter(filename, mode="a", engine="openpyxl")   # open agagin and all lost
# Note that creating an ExcelWriter object with a file name that already exists will result in the contents of the existing file being erased.
existing = df
df2 = pd.concat([df2, existing])
df2.to_excel(writerBen, sheet_name='sheet2')
writerBen._save()

writerBen = pd.ExcelWriter(filename, mode="a", engine="openpyxl")   # open agagin and all los

workbook = openpyxl.load_workbook(filename)
# Check if the sheet exists before attempting to delete
if 'sheet2' in workbook.sheetnames:
    # Delete the specified sheet
    workbook.remove(workbook['sheet2'])
    workbook.save(filename)
workbook.close()

with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writerBen:
    df2.to_excel(writerBen, sheet_name='Sheet2', index=False)

writerBen = pd.ExcelWriter(filename, mode="a", engine="openpyxl")   # open agagin and all lost
# with pd.ExcelWriter(filename, mode="a", engine="openpyxl") as writerBen:
df.to_excel(writerBen, sheet_name="Sheet3")
writerBen._save()

# file = pd.ExcelFile('myfile.xlsx')
# file.parse()

excel_file_path = filename
workbook = openpyxl.load_workbook(excel_file_path)
if 'Sheet2' in workbook.sheetnames:
    # Delete the specified sheet
    workbook.remove(workbook['Sheet2'])
    workbook.save(excel_file_path)
workbook.close()

writerBen= pd.ExcelWriter(excel_file_path, mode="a", engine="openpyxl")
if not df2.empty:
    df2.to_excel(writerBen, sheet_name="Sheet2", index=False)
else:
    print("DataFrame is empty. Not writing to the Excel file.")
writerBen._save()
writer= pd.ExcelWriter(excel_file_path, mode="a", engine="openpyxl")
    # Write the DataFrame to Excel with adjusted column width
df.to_excel(writer, sheet_name='Sheet6',index=False)

# Access the worksheet
with pd.ExcelWriter(excel_file_path, mode="a", engine="openpyxl") as writer:
    # Write the DataFrame to Sheet1
    # df.to_excel(writer, sheet_name='Sheet3', index=False)

    # Access the worksheet for Sheet1
    worksheet1 = writer.sheets['Sheet1']

    # Adjust the column width for Column1 in Sheet1
    worksheet1.column_dimensions['A'].width = 30

    # Access the worksheet for Sheet2
    worksheet2 = writer.sheets['Sheet2']

    # Adjust the column width for ColumnA in Sheet2
    worksheet2.column_dimensions['A'].width = 20

# worksheet.save(filename)
